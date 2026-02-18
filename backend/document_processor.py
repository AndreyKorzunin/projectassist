import os
import re
from typing import Dict, Any
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
import pandas as pd
from PyPDF2 import PdfReader
import structlog

logger = structlog.get_logger()


class DocumentProcessor:


    def process(self, file_path: str) -> Dict[str, Any]:

        _, ext = os.path.splitext(file_path)

        if ext == '.docx':
            return self.process_word(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self.process_excel(file_path)
        elif ext == '.pdf':
            return self.process_pdf(file_path)
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")

    def process_word(self, file_path: str) -> Dict[str, Any]:

        logger.info("processing_word", file=file_path)
        doc = Document(file_path)
        content = {
            "metadata": {
                "filename": os.path.basename(file_path),
                "type": "word",
                "extension": ".docx"
            },
            "headings": [],
            "paragraphs": [],
            "tables": [],
            "lists": [],
            "full_text": ""
        }

        current_list = None

        for element in doc.element.body:
            if element.tag.endswith('tbl'):
                table = Table(element, doc)
                table_data = self._extract_table_data(table)
                if table_data["rows"] > 0:
                    content["tables"].append(table_data)

            elif element.tag.endswith('p'):
                paragraph = Paragraph(element, doc)
                text = paragraph.text.strip()
                if not text:
                    continue

                content["full_text"] += text + "\n"


                if paragraph.style and paragraph.style.name.startswith('Heading'):
                    level_match = re.search(r'\d+', paragraph.style.name)
                    level = int(level_match.group()) if level_match else 1
                    content["headings"].append({
                        "level": level,
                        "text": text,
                        "index": len(content["paragraphs"])
                    })


                elif paragraph.style and paragraph.style.name in ['List Paragraph', 'ListBullet', 'ListNumber']:
                    if current_list is None:
                        list_type = "bullet" if "Bullet" in paragraph.style.name else "number"
                        current_list = {"type": list_type, "items": []}
                    current_list["items"].append(text)

                else:
                    if current_list and current_list["items"]:
                        content["lists"].append(current_list)
                        current_list = None
                    content["paragraphs"].append(text)


        if current_list and current_list["items"]:
            content["lists"].append(current_list)


        content["statistics"] = {
            "total_words": len(content["full_text"].split()),
            "total_characters": len(content["full_text"]),
            "headings_count": len(content["headings"]),
            "paragraphs_count": len(content["paragraphs"]),
            "tables_count": len(content["tables"]),
            "lists_count": len(content["lists"])
        }

        logger.info("word_processed", stats=content["statistics"])
        return content

    def _extract_table_data(self, table: Table) -> Dict[str, Any]:

        data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                row_data.append(cell.text.strip())
            if any(row_data):
                data.append(row_data)

        return {
            "rows": len(data),
            "cols": len(data[0]) if data else 0,
            "data": data,
            "has_header": self._detect_header(data)
        }

    def _detect_header(self, data: List[List[str]]) -> bool:
        
        if len(data) < 2:
            return False

        header_row = data[0]


        if not header_row or all(cell.strip() == "" for cell in header_row):
            return False

        if not all(len(cell) < 50 for cell in header_row):
            return False

        for cell in header_row[:3]:
            cleaned = cell.replace('.', '').replace(',', '').strip()
            if cleaned.isdigit() or (cleaned and all(c.isdigit() or c in ('.', ',') for c in cleaned)):
                return False


        if len(data) > 1:
            data_row = data[1]

            if len(header_row) == len(data_row):
                similarity = sum(1 for h, d in zip(header_row, data_row) if h.strip() == d.strip())
                if similarity / len(header_row) > 0.7:
                    return False

        return True

def process_excel(self, file_path: str) -> Dict[str, Any]:

    logger.info("processing_excel", file=file_path)
    wb = load_workbook(file_path, data_only=True)
    content = {
        "metadata": {
            "filename": os.path.basename(file_path),
            "type": "excel",
            "extension": os.path.splitext(file_path)[1]
        },
        "sheets": {},
        "full_text": ""
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data_rows = []

        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append([str(cell) if cell is not None else "" for cell in row])

        if not data_rows:
            continue

        df = pd.DataFrame(data_rows)


        headers = None
        if len(df) > 1:
            first_row = df.iloc[0]
            if all(isinstance(val, str) and len(str(val)) < 50 for val in first_row):
                headers = first_row.tolist()
                df = df[1:]

        sheet_data = {
            "rows": len(df),
            "cols": len(df.columns),
            "headers": headers,
            "data_preview": df.head(10).values.tolist() if not df.empty else [],
            "numeric_columns": self._detect_numeric_columns(df)
        }

        content["sheets"][sheet_name] = sheet_data

        sheet_text = f"Лист: {sheet_name}\n"
        if headers:
            sheet_text += "Заголовки: " + ", ".join(str(h) for h in headers if h) + "\n"
        content["full_text"] += sheet_text + "\n"


    total_rows = sum(sheet["rows"] for sheet in content["sheets"].values())
    content["statistics"] = {
        "sheets_count": len(content["sheets"]),
        "total_rows": total_rows,
        "total_sheets": len(content["sheets"])
    }

    logger.info("excel_processed", sheets=len(content["sheets"]))
    return content


def _detect_numeric_columns(self, df: pd.DataFrame) -> list:

    numeric_cols = []
    for col_idx, col in enumerate(df.columns):
        sample = df[col].dropna().head(10)
        if len(sample) > 0:
            numeric_count = sum(
                isinstance(x, (int, float)) or
                (isinstance(x, str) and re.match(r'^-?\d+(\.\d+)?$', str(x).strip()))
                for x in sample
            )
            if numeric_count / len(sample) > 0.7:
                numeric_cols.append(col_idx)
    return numeric_cols


def process_pdf(self, file_path: str) -> Dict[str, Any]:

    logger.info("processing_pdf", file=file_path)
    reader = PdfReader(file_path)
    full_text = ""

    for page_num, page in enumerate(reader.pages, 1):
        text = page.extract_text()
        if text:
            full_text += text + "\n"

    content = {
        "metadata": {
            "filename": os.path.basename(file_path),
            "type": "pdf",
            "extension": ".pdf",
            "pages": len(reader.pages)
        },
        "full_text": full_text,
        "statistics": {
            "total_words": len(full_text.split()),
            "total_characters": len(full_text),
            "pages": len(reader.pages)
        }
    }

    logger.info("pdf_processed", pages=len(reader.pages))
    return content